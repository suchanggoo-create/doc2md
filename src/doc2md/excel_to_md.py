from __future__ import annotations

import html
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class CellSpan:
    rowspan: int = 1
    colspan: int = 1


def _sheet_used_bounds(ws: Worksheet) -> Tuple[int, int, int, int]:
    """
    Return (min_row, min_col, max_row, max_col) for the sheet's used range.
    Uses calculate_dimension() which is more reliable for "formatted" sheets than max_row/max_col.
    """
    dim = ws.calculate_dimension()  # like "A1:D20" (or "A1" for empty-ish)
    if ":" in dim:
        start, end = dim.split(":", 1)
    else:
        start, end = dim, dim
    min_row = ws[start].row
    min_col = ws[start].column
    max_row = ws[end].row
    max_col = ws[end].column

    # Expand by merged ranges that extend beyond calculate_dimension.
    for r in ws.merged_cells.ranges:
        min_row = min(min_row, r.min_row)
        min_col = min(min_col, r.min_col)
        max_row = max(max_row, r.max_row)
        max_col = max(max_col, r.max_col)
    return min_row, min_col, max_row, max_col


def _merged_span_maps(ws: Worksheet) -> Tuple[Dict[Tuple[int, int], CellSpan], set[Tuple[int, int]]]:
    """
    Returns:
    - span_by_anchor[(r,c)] = CellSpan(rowspan, colspan) for the top-left anchor cell of a merged range.
    - covered_cells = set of (r,c) that are inside a merged range but are NOT the anchor.
    """
    span_by_anchor: Dict[Tuple[int, int], CellSpan] = {}
    covered: set[Tuple[int, int]] = set()
    for r in ws.merged_cells.ranges:
        anchor = (r.min_row, r.min_col)
        span_by_anchor[anchor] = CellSpan(rowspan=r.max_row - r.min_row + 1, colspan=r.max_col - r.min_col + 1)
        for rr in range(r.min_row, r.max_row + 1):
            for cc in range(r.min_col, r.max_col + 1):
                if (rr, cc) != anchor:
                    covered.add((rr, cc))
    return span_by_anchor, covered


def render_sheet_as_html_table(ws: Worksheet, *, max_rows: Optional[int] = None) -> str:
    """
    Render a sheet as an HTML table with rowspan/colspan to preserve merged cells.
    max_rows limits output height (starting from min_row).
    """
    min_row, min_col, max_row, max_col = _sheet_used_bounds(ws)
    if max_rows is not None:
        max_row = min(max_row, min_row + max_rows - 1)

    span_by_anchor, covered = _merged_span_maps(ws)

    lines: List[str] = ["<table>"]
    for r in range(min_row, max_row + 1):
        lines.append("  <tr>")
        for c in range(min_col, max_col + 1):
            if (r, c) in covered:
                continue
            span = span_by_anchor.get((r, c), CellSpan())
            v = ws.cell(row=r, column=c).value
            text = "" if v is None else str(v).replace("\n", " ")
            text = html.escape(text)

            attrs = ""
            if span.rowspan > 1:
                attrs += f' rowspan="{span.rowspan}"'
            if span.colspan > 1:
                attrs += f' colspan="{span.colspan}"'
            lines.append(f"    <td{attrs}>{text}</td>")
        lines.append("  </tr>")
    lines.append("</table>")
    return "\n".join(lines)


def render_xlsx_file_to_markdown(xlsx_path: Path, *, preview_rows: int = 40) -> str:
    """
    Render a full xlsx file to Markdown, one sheet at a time.
    Uses HTML <table> to preserve merged cells.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)
    try:
        out: List[str] = []
        for sheet in wb.worksheets:
            out.append(f"## {sheet.title}")
            out.append("")
            out.append(render_sheet_as_html_table(sheet, max_rows=None if preview_rows <= 0 else preview_rows))
            out.append("")
            if preview_rows > 0:
                # Best-effort: if the sheet has more rows than rendered, annotate truncation.
                min_row, _, max_row, _ = _sheet_used_bounds(sheet)
                total = max_row - min_row + 1
                if total > preview_rows:
                    out.append(f"> [TRUNCATED: showing first {preview_rows} of {total} rows]")
                    out.append("")
        return "\n".join(out).rstrip() + "\n"
    finally:
        wb.close()


def write_xlsx_as_markdown(xlsx_path: Path, md_path: Path, *, preview_rows: int = 40) -> None:
    md = render_xlsx_file_to_markdown(xlsx_path, preview_rows=preview_rows)
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(md, encoding="utf-8")

